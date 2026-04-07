

import re
import pdfplumber
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Config ─────────────────────────────────────────────────────────────────────
PDF_PATH  = r"C:\Users\SamuelJoshuaRaj\Downloads\AC119.pdf"
XLSX_PATH = r"C:\Users\SamuelJoshuaRaj\OneDrive - CYGNUSA Technologies\Desktop\New folder\AC119_Results.xlsx"

# ── Column headers (exactly matching ECI Form-20) ─────────────────────────────
CANDIDATE_HEADERS = [
    "Sl. No.",
    "Polling Station No.",
    "Karthikeya Sivasenapathy\n(DMK)",
    "Velumani S.P.\n(AIADMK)",
    "Kalaiarasi S.\n(NTK)",
    "Sathish Kumar S.R.\n(AMMK)",
    "Badran M.\n(Ganasangam Party of India)",
    "Shajahan S.\n(MNM)",
    "Abdul Gafoor A.\n(Independent)",
    "Selvamohan C.C.\n(Independent)",
    "Mansoor Alikhan A.\n(Independent)",
    "John Edward Visuvasam J.\n(Independent)",
    "NOTA",
    "Total of Valid Votes",
    "No. of Rejected Votes",
    "No. of Tendered Votes",
]

TOTAL_COLS = len(CANDIDATE_HEADERS)   # 16


def parse_rows_from_pdf(pdf_path: str):
    """
    Walk every page and extract numeric data rows.

    Each Form-20 row has exactly 17 whitespace-separated tokens:
        sl_no  ps_no  c1 c2 c3 c4 c5 c6 c7 c8 c9 c10  nota  total_valid  rejected  total_polled  tendered
    We keep all 17 but map them to our 16-column layout:
        [sl, ps, c1..c10, nota, total_valid, rejected, tendered]
    (total_polled == total_valid for valid rows, so we drop it.)
    """
    rows = []
    seen_sl = set()   # de-duplicate across pages

    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            text = page.extract_text()
            if not text:
                continue

            for line in text.splitlines():
                tokens = line.split()
                if not tokens:
                    continue

                # First token must be a serial number (integer 1-600)
                if not tokens[0].isdigit():
                    continue
                sl_no = int(tokens[0])
                if not (1 <= sl_no <= 600):
                    continue
                if sl_no in seen_sl:
                    continue

                # Need at least 17 tokens total
                if len(tokens) < 17:
                    continue

                ps_no = tokens[1]    # e.g. "1", "3A", "11A"

                # tokens[2..16] are the 15 numeric fields:
                # c1 c2 c3 c4 c5 c6 c7 c8 c9 c10 nota total_valid rejected total_polled tendered
                try:
                    nums = [int(t) for t in tokens[2:17]]
                except ValueError:
                    continue

                if len(nums) != 15:
                    continue

                # Map to our 16-col layout (drop total_polled at index 13):
                row = (
                    [sl_no, ps_no]
                    + nums[0:11]       # c1..c10, nota
                    + [nums[11]]       # total_valid
                    + [nums[12]]       # rejected
                    + [nums[14]]       # tendered  (skip nums[13] = total_polled)
                )
                rows.append(row)
                seen_sl.add(sl_no)

    # Sort by serial number
    rows.sort(key=lambda r: r[0])
    return rows


# Summary rows hardcoded from the last page of the PDF
SUMMARY_ROWS = [
    ["Polling Station Totals",
     81829, 123538, 8014, 1229, 136, 11572, 141, 173, 426, 311, 1622, 228991, 0, 0],
    ["Postal Ballot Totals",
     766, 687, 28, 6, 0, 34, 1, 2, 2, 1, 13, 1540, 883, 0],
    ["Grand Total (Votes Polled)",
     82595, 124225, 8042, 1235, 136, 11606, 142, 175, 428, 312, 1635, 230531, 883, 0],
]


def write_excel(rows: list, xlsx_path: str) -> None:
    """Write extracted rows to a formatted Excel workbook."""

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Form-20 Results"

    # ── Styles ────────────────────────────────────────────────────────────────
    HEADER_FILL  = PatternFill("solid", fgColor="1A3A5C")
    ALT_FILL     = PatternFill("solid", fgColor="EBF3FB")
    WHITE_FILL   = PatternFill("solid", fgColor="FFFFFF")
    SUM_FILL     = PatternFill("solid", fgColor="FFF2CC")

    HEADER_FONT  = Font(name="Arial", bold=True, color="FFFFFF", size=9)
    DATA_FONT    = Font(name="Arial", size=9)
    SUM_FONT     = Font(name="Arial", bold=True, size=9)
    TITLE_FONT   = Font(name="Arial", bold=True, color="FFFFFF", size=10)

    thin   = Side(style="thin", color="BBBBBB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left   = Alignment(horizontal="left",   vertical="center", wrap_text=True)

    # ── Row 1: Title ──────────────────────────────────────────────────────────
    ws.merge_cells(1, 1, 1, TOTAL_COLS)
    c = ws.cell(row=1, column=1)
    c.value = (
        "FORM-20 — FINAL RESULT SHEET  |  119 Thondamuthur  |  "
        "General Elections to Assembly Constituency 2021  |  "
        "Total Electors: 3,26,799  |  Date: 02-05-2021"
    )
    c.font      = TITLE_FONT
    c.fill      = HEADER_FILL
    c.alignment = center
    ws.row_dimensions[1].height = 22

    # ── Row 2: Column headers ─────────────────────────────────────────────────
    for ci, hdr in enumerate(CANDIDATE_HEADERS, 1):
        c = ws.cell(row=2, column=ci)
        c.value     = hdr
        c.font      = HEADER_FONT
        c.fill      = HEADER_FILL
        c.alignment = center
        c.border    = border
    ws.row_dimensions[2].height = 46

    # ── Data rows ─────────────────────────────────────────────────────────────
    for ri, row_data in enumerate(rows, start=3):
        fill = ALT_FILL if ri % 2 == 0 else WHITE_FILL
        for ci, val in enumerate(row_data, 1):
            c = ws.cell(row=ri, column=ci)
            c.value     = val
            c.font      = DATA_FONT
            c.fill      = fill
            c.border    = border
            c.alignment = left if ci == 2 else center
        ws.row_dimensions[ri].height = 14

    # ── Blank separator ───────────────────────────────────────────────────────
    sep = len(rows) + 3
    ws.row_dimensions[sep].height = 8

    # ── Summary rows ─────────────────────────────────────────────────────────
    for si, (label, *vals) in enumerate(SUMMARY_ROWS):
        er = sep + 1 + si
        ws.cell(row=er, column=1).value = ""
        ws.cell(row=er, column=1).fill  = SUM_FILL
        ws.cell(row=er, column=1).border= border
        ws.cell(row=er, column=1).font  = SUM_FONT

        c2 = ws.cell(row=er, column=2)
        c2.value     = label
        c2.font      = SUM_FONT
        c2.fill      = SUM_FILL
        c2.border    = border
        c2.alignment = left

        for ci, v in enumerate(vals, 3):
            c = ws.cell(row=er, column=ci)
            c.value     = v if v != "" else None
            c.font      = SUM_FONT
            c.fill      = SUM_FILL
            c.border    = border
            c.alignment = center
        ws.row_dimensions[er].height = 16

    # ── Column widths ─────────────────────────────────────────────────────────
    widths = [6, 9, 14, 12, 10, 13, 17, 11, 13, 13, 14, 17, 8, 14, 12, 12]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Freeze header rows
    ws.freeze_panes = "C3"

    wb.save(xlsx_path)
    print(f"Saved → {xlsx_path}  ({len(rows)} data rows + {len(SUMMARY_ROWS)} summary rows)")


if __name__ == "__main__":
    print(f"Reading: {PDF_PATH}")
    rows = parse_rows_from_pdf(PDF_PATH)
    print(f"  Extracted {len(rows)} polling-station rows")
    write_excel(rows, XLSX_PATH)