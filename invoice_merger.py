"""
Invoice Merger — March Invoice → Rental Desktop & Laptop Configuration
=======================================================================
Dependencies:
    pip install pandas openpyxl xlrd

Usage:
    1. Place march_invoice.xls and Rental_Desktop_and_Laptop_Configurations.xlsx
       in the same folder as this script  (or update the paths below).
    2. Run:  python invoice_merger.py
    3. Output: Rental_Desktop_Laptop_Updated.xlsx  (same folder)

What this script does
─────────────────────
• Reads the March invoice (.xls) and the rental configuration (.xlsx).
• Matches every invoice line to a row in the "Date of Delivery" sheet
  using the primary serial number AND the back-door serial number.
• Fills in missing Balaji Info Cost (column W) from the invoice Amount.
• Marks doubtful / mismatched cells in LIGHT PURPLE (CC99FF):
    – Balaji cost mismatch  (rental value ≠ invoice amount by > ₹5)
    – Serial number issues  (monitor serial used for desktop row, SN typos)
    – Date mismatch         (delivery date differs between files)
• Appends notes in the Remarks column explaining each purple flag.
• Adds 4 brand-new rows for invoice lines that had no matching entry:
    S.No 96 – MacBook Pro A1990      (C02Z310HLVDT)
    S.No 97 – HP Elite Book 840 G6   (5CG0165ZBM)
    S.No 98 – Dell Pro PV14250       (3FVZYC4)
    S.No 99 – Dell Latitude 5400     (HSL4Y33)
• Preserves ALL existing cell styles, fonts, borders, and formatting.
"""

import copy
import datetime
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# ── File paths ────────────────────────────────────────────────────────────────
INVOICE_PATH = r"C:\Users\SamuelJoshuaRaj\Downloads\BIS_March_invocie\march invoice.xls"
RENTAL_PATH  = r"C:\Users\SamuelJoshuaRaj\Downloads\BIS_March_invocie\Rental Desktop and Laptop Configurations.xlsx"
OUTPUT_PATH  = r"C:\Users\SamuelJoshuaRaj\OneDrive - CYGNUSA Technologies\Desktop\Balaji invoices\Rental_Desktop_Laptop_Updated.xlsx"

# ── Highlight colour ──────────────────────────────────────────────────────────
LIGHT_PURPLE = PatternFill(start_color="CC99FF", end_color="CC99FF", fill_type="solid")

# ── Excel column indices (1-based) for "Date of Delivery" sheet ───────────────
COL_SN           = 1   # A  – S.No
COL_DC           = 2   # B  – DC Number
COL_DATE         = 3   # C  – Date of Delivered
COL_USER         = 4   # D  – User Name
COL_SPEC         = 5   # E  – Specification
COL_ASSET        = 6   # F  – Asset Type
COL_MAKE         = 7   # G  – Make
COL_QTY          = 8   # H  – Quantity
COL_ADAPTER      = 9   # I  – Adapter
COL_BAG          = 10  # J  – Bag
COL_LAPTOP       = 11  # K  – Laptop (model name / description)
COL_CONFIG       = 12  # L  – Configuration
COL_SERIAL       = 13  # M  – Serial Number
COL_HDMI         = 14  # N  – HDMI Cable
COL_POWER        = 15  # O  – Power Cable
COL_KEYBOARD     = 16  # P  – Keyboard
COL_MOUSE        = 17  # Q  – Mouse
COL_MONITOR      = 18  # R  – Monitor
COL_MEDCOST      = 19  # S  – Mediantlabs Cost per Month
COL_STATUS       = 20  # T  – Status
COL_RETDC        = 21  # U  – Return DC Number
COL_RETDATE      = 22  # V  – Date of return
COL_BALAJI_COST  = 23  # W  – Balaji Info Cost
COL_DAYS         = 24  # X  – Number of Days
COL_MEDANALYSIS  = 25  # Y  – Mediantlabs Cost analysis
COL_BALAJI_ANAL  = 26  # Z  – Balaji Cost Analysis
# col 27 is blank (AA)
COL_REMARKS      = 28  # AB – Remarks


# ═══════════════════════════════════════════════════════════════════════════════
# HELPERS
# ═══════════════════════════════════════════════════════════════════════════════

def clean_sn(value):
    """Strip and upper-case a serial number string; return '' for nulls."""
    if pd.isna(value):
        return ""
    return str(value).strip().upper()


def parse_date(value):
    """
    Try to convert various date representations to a datetime object.
    Accepts: datetime, 'dd.mm.yyyy', 'yyyy-mm-dd', 'dd/mm/yyyy'.
    Returns the original value unchanged if no format matches.
    """
    if pd.isna(value):
        return None
    if isinstance(value, (datetime.datetime, datetime.date)):
        return value
    s = str(value).strip()
    for fmt in ("%d.%m.%Y", "%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y"):
        try:
            return datetime.datetime.strptime(s, fmt)
        except ValueError:
            pass
    return s  # leave as-is if unparseable


def copy_cell_style(src, dst):
    """
    Copy font, border, alignment, fill, and number-format from src to dst.
    Does NOT copy the value.
    """
    if src.has_style:
        dst.font         = copy.copy(src.font)
        dst.border       = copy.copy(src.border)
        dst.alignment    = copy.copy(src.alignment)
        dst.fill         = copy.copy(src.fill)
        dst.number_format = src.number_format


def append_remark(ws, excel_row, note):
    """Append a note to the Remarks cell, avoiding duplicate entries."""
    cell = ws.cell(row=excel_row, column=COL_REMARKS)
    existing = str(cell.value) if cell.value not in (None, "nan", "None") else ""
    if note not in existing:
        cell.value = (existing + " " + note).strip() if existing else note


def mark_purple(ws, excel_row, col):
    """Apply light-purple fill to a single cell."""
    ws.cell(row=excel_row, column=col).fill = LIGHT_PURPLE


# ═══════════════════════════════════════════════════════════════════════════════
# STEP 1 — Load the March invoice
# ═══════════════════════════════════════════════════════════════════════════════

print("Loading march invoice …")
raw = pd.read_excel(INVOICE_PATH, sheet_name="Sheet1", header=None)

# Row 1  (index 1) = column headers
# Rows 2–75 (index 2–75) = invoice line items
inv_data = raw.iloc[2:76].copy().reset_index(drop=True)
inv_data.columns = range(len(inv_data.columns))

# Invoice column layout:
#  0 = No.          1 = D.C.No       2 = DC Date       3 = Description
#  4 = MODEL        5 = Serial No.   6 = BACK DOOR NO  7 = CONFIG
#  8 = Period       9 = Per Day (Rs) 10= Day           11= Amount (Rs.)
# 12 = Return Dc   13 = Return Date  14= TALLY DC NO   15= PERSON
# 16 = QTY         17= Remarks

# Build  serial → row-index  lookup (covers primary AND back-door serials)
inv_by_serial = {}   # key: cleaned SN  →  value: index into inv_data
inv_row_indices = [] # ordered list of valid invoice row indices

for i, row in inv_data.iterrows():
    if pd.isna(row[0]):          # skip blank rows
        continue
    inv_row_indices.append(i)
    sn = clean_sn(row[5])
    bd = clean_sn(row[6])
    if sn:
        inv_by_serial[sn] = i
    if bd and bd != sn:
        inv_by_serial[bd] = i

print(f"  Invoice lines loaded: {len(inv_row_indices)}")


def get_inv_idx(serial_field_raw):
    """
    Given the raw value of a rental-config Serial Number cell
    (which may be 'SN1 / SN2'), return the matching invoice row index or None.
    """
    parts = [p.strip().upper() for p in str(serial_field_raw).split("/")]
    for p in parts:
        if p in inv_by_serial:
            return inv_by_serial[p]
    return None


def invoice_amount(inv_idx):
    """Return the invoice Amount (Rs.) for a given invoice row index, or None."""
    amt = inv_data.iloc[inv_idx][11]
    if pd.notna(amt):
        try:
            v = float(amt)
            return v if v > 0 else None
        except (ValueError, TypeError):
            pass
    return None


# ═══════════════════════════════════════════════════════════════════════════════
# STEP 2 — Load the rental configuration workbook (preserving all styles)
# ═══════════════════════════════════════════════════════════════════════════════

print("Loading rental configuration workbook …")
wb = load_workbook(RENTAL_PATH)
ws = wb["Date of Delivery"]

# Identify the last row that actually has an S.No value
last_data_row = 1
for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    if row[COL_SN - 1].value not in (None, "", "nan"):
        last_data_row = row[COL_SN - 1].row

print(f"  Last data row in rental sheet: {last_data_row}")

# Build  cleaned-serial → excel-row  lookup for the rental sheet
rental_serial_to_row = {}
for row in ws.iter_rows(min_row=2, max_row=last_data_row):
    sno = row[COL_SN - 1].value
    if sno in (None, "", "nan"):
        continue
    excel_row = row[COL_SN - 1].row
    raw_sn = str(ws.cell(row=excel_row, column=COL_SERIAL).value or "")
    for part in raw_sn.split("/"):
        p = part.strip().upper()
        if p and p not in ("NAN", "NOT APPLICABLE", ""):
            rental_serial_to_row[p] = excel_row


# ═══════════════════════════════════════════════════════════════════════════════
# STEP 3 — Pass 1: update existing rental rows
# ═══════════════════════════════════════════════════════════════════════════════

print("\nPass 1 — updating matched rows …")
matched_inv_indices = set()

for row in ws.iter_rows(min_row=2, max_row=last_data_row):
    sno_val = row[COL_SN - 1].value
    if sno_val in (None, "", "nan"):
        continue

    excel_row   = row[COL_SN - 1].row
    serial_raw  = ws.cell(row=excel_row, column=COL_SERIAL).value
    if serial_raw is None:
        continue

    inv_idx = get_inv_idx(str(serial_raw))
    if inv_idx is None:
        continue  # no invoice match for this rental row

    matched_inv_indices.add(inv_idx)
    inv_row    = inv_data.iloc[inv_idx]
    inv_amt    = invoice_amount(inv_idx)
    balaji_cell = ws.cell(row=excel_row, column=COL_BALAJI_COST)
    current_val = balaji_cell.value

    # ── Fill missing Balaji Info Cost ──────────────────────────────────────
    if current_val in (None, "", "nan") and inv_amt is not None:
        balaji_cell.value = round(inv_amt, 2)
        print(f"  S.No {sno_val:>3}: filled  Balaji Cost = {inv_amt:.2f}  ({serial_raw})")

    # ── Flag cost mismatch ─────────────────────────────────────────────────
    elif current_val not in (None, "", "nan") and inv_amt is not None:
        try:
            if abs(float(current_val) - inv_amt) > 5:
                mark_purple(ws, excel_row, COL_BALAJI_COST)
                note = (f"[MISMATCH: Balaji cost in rental={current_val} "
                        f"vs invoice amount={inv_amt:.2f}]")
                append_remark(ws, excel_row, note)
                print(f"  S.No {sno_val:>3}: MISMATCH  rental={current_val}  "
                      f"invoice={inv_amt:.2f}  ({serial_raw})")
        except (ValueError, TypeError):
            pass

    # ── Flag back-door serial used as primary (info note) ─────────────────
    primary_sn  = clean_sn(inv_row[5])
    backdoor_sn = clean_sn(inv_row[6])
    rental_parts = [p.strip().upper() for p in str(serial_raw).split("/")]

    if (backdoor_sn
            and backdoor_sn in rental_parts
            and primary_sn
            and primary_sn not in rental_parts):
        append_remark(ws, excel_row, f"[Invoice primary serial: {primary_sn}]")


# ── Special: S.No 1 & 2 — monitor serials used for desktop rows ────────────
for target_sno in (1, 2):
    desktop_sn = "BIS-053" if target_sno == 1 else "BIS-054"
    for row in ws.iter_rows(min_row=2, max_row=last_data_row):
        if row[COL_SN - 1].value == target_sno:
            er = row[COL_SN - 1].row
            mark_purple(ws, er, COL_SERIAL)
            append_remark(ws, er,
                f"[DOUBTFUL: Monitor serial used for Desktop row. "
                f"Invoice desktop serial: {desktop_sn}]")
            break

# ── Special: S.No 76 — back-door serial typo (3684N13 vs 3694N13) ──────────
for row in ws.iter_rows(min_row=2, max_row=last_data_row):
    if row[COL_SN - 1].value == 76:
        er = row[COL_SN - 1].row
        sn_val = str(ws.cell(row=er, column=COL_SERIAL).value or "")
        if "3684N13" in sn_val.upper():
            mark_purple(ws, er, COL_SERIAL)
            append_remark(ws, er,
                "[MISMATCH: Invoice back-door serial is 3694N13 but rental shows 3684N13]")
        break

# ── Special: S.No 83 — date mismatch (rental=10.03.2026, invoice=13.03.2026) ─
for row in ws.iter_rows(min_row=2, max_row=last_data_row):
    if row[COL_SN - 1].value == 83:
        er = row[COL_SN - 1].row
        mark_purple(ws, er, COL_DATE)
        append_remark(ws, er,
            "[MISMATCH: Date in rental=10.03.2026, Invoice DC date=13.03.2026]")
        break


# ═══════════════════════════════════════════════════════════════════════════════
# STEP 4 — Identify invoice lines with no rental match
# ═══════════════════════════════════════════════════════════════════════════════

print("\nIdentifying unmatched invoice lines …")

# Invoice lines to skip even if unmatched:
#   • BIS-053 / BIS-054 → already represented under monitor serials (S.No 1 & 2)
#   • Monitors (Model='Monitor', Amount=0) → not tracked as separate rows
SKIP_SERIALS = {"BIS - 053", "BIS-053", "BIS - 054", "BIS-054"}

rows_to_add = []
for i in inv_row_indices:
    if i in matched_inv_indices:
        continue
    row = inv_data.iloc[i]
    sn  = clean_sn(row[5])
    if sn in SKIP_SERIALS:
        continue
    model = str(row[4]).strip().upper() if pd.notna(row[4]) else ""
    amt   = invoice_amount(i)
    if model == "MONITOR" and (amt is None or amt == 0):
        continue
    rows_to_add.append(i)
    print(f"  Will add → Invoice No {row[0]:>3}: serial={row[5]}  DC={row[1]}  {row[3]}")

print(f"  Rows to add: {len(rows_to_add)}")


# ═══════════════════════════════════════════════════════════════════════════════
# STEP 5 — Pass 2: append missing invoice lines as new rows
# ═══════════════════════════════════════════════════════════════════════════════

print("\nPass 2 — appending new rows …")

# Find the current maximum S.No so we can continue numbering
current_max_sno = 0
for row in ws.iter_rows(min_row=2, max_row=last_data_row):
    v = row[COL_SN - 1].value
    if v is not None:
        try:
            current_max_sno = max(current_max_sno, int(v))
        except (ValueError, TypeError):
            pass

# Use the second-to-last data row as a style template (avoids any anomalies
# at the very last row such as partial entries or different fill colours).
template_row = last_data_row - 1

for offset, inv_idx in enumerate(rows_to_add):
    inv_row      = inv_data.iloc[inv_idx]
    new_sno      = current_max_sno + offset + 1
    new_excel_row = last_data_row + offset + 1

    # ── Copy styles from template row ─────────────────────────────────────
    for col in range(1, 29):
        src = ws.cell(row=template_row, column=col)
        dst = ws.cell(row=new_excel_row, column=col)
        copy_cell_style(src, dst)
        dst.value = None  # clear value; we fill selectively below

    # ── Derive make / asset type from description ──────────────────────────
    model = str(inv_row[4]).strip().upper() if pd.notna(inv_row[4]) else ""
    desc  = str(inv_row[3]).strip()         if pd.notna(inv_row[3]) else ""

    if model == "MACBOOK":
        asset_type, make, spec = "Laptop", "Apple", "Business"
    elif model == "LAPTOP":
        asset_type, spec = "Laptop", "Business"
        desc_upper = desc.upper()
        if "HP" in desc_upper:
            make = "HP"
        elif "DELL" in desc_upper:
            make = "Dell"
        elif "LENOVO" in desc_upper:
            make = "Lenovo"
        else:
            make = ""
    else:
        asset_type, make, spec = model, "", model

    # ── Serial number (combine primary + back-door if different) ──────────
    sn  = str(inv_row[5]).strip() if pd.notna(inv_row[5]) else ""
    bd  = str(inv_row[6]).strip() if pd.notna(inv_row[6]) else ""
    serial_str = (f"{sn} / {bd}"
                  if bd and bd.upper() != sn.upper() and bd.lower() not in ("nan", "")
                  else sn)

    # ── Status + return info ───────────────────────────────────────────────
    has_return = (pd.notna(inv_row[12])
                  and str(inv_row[12]).strip() not in ("", "nan"))
    status     = "Returned" if has_return else "Active"

    # ── Remarks ───────────────────────────────────────────────────────────
    inv_remark = (str(inv_row[17]).strip()
                  if pd.notna(inv_row[17]) and str(inv_row[17]).strip().lower() != "nan"
                  else "")
    remark_text = ("[Added from March Invoice] " + inv_remark).strip()

    # ── Write cells ───────────────────────────────────────────────────────
    ws.cell(row=new_excel_row, column=COL_SN).value           = new_sno
    ws.cell(row=new_excel_row, column=COL_DC).value           = str(inv_row[1]).strip()
    ws.cell(row=new_excel_row, column=COL_DATE).value         = parse_date(inv_row[2])
    ws.cell(row=new_excel_row, column=COL_USER).value         = "Unknown"
    ws.cell(row=new_excel_row, column=COL_SPEC).value         = spec
    ws.cell(row=new_excel_row, column=COL_ASSET).value        = asset_type
    ws.cell(row=new_excel_row, column=COL_MAKE).value         = make
    ws.cell(row=new_excel_row, column=COL_QTY).value          = 1
    ws.cell(row=new_excel_row, column=COL_ADAPTER).value      = 1
    ws.cell(row=new_excel_row, column=COL_BAG).value          = 1
    ws.cell(row=new_excel_row, column=COL_LAPTOP).value       = desc
    ws.cell(row=new_excel_row, column=COL_CONFIG).value       = (
        str(inv_row[7]).strip() if pd.notna(inv_row[7]) else "")
    ws.cell(row=new_excel_row, column=COL_SERIAL).value       = serial_str
    ws.cell(row=new_excel_row, column=COL_STATUS).value       = status
    ws.cell(row=new_excel_row, column=COL_REMARKS).value      = remark_text

    # Accessories — standard default for laptops
    for col in (COL_HDMI, COL_POWER, COL_KEYBOARD, COL_MOUSE, COL_MONITOR):
        ws.cell(row=new_excel_row, column=col).value = "Not Applicable"

    # Balaji Info Cost = invoice Amount
    amt = invoice_amount(inv_idx)
    ws.cell(row=new_excel_row, column=COL_BALAJI_COST).value = (
        round(amt, 2) if amt is not None else None)

    # Number of Days
    days = inv_row[10]
    ws.cell(row=new_excel_row, column=COL_DAYS).value = (
        int(days) if pd.notna(days) else None)

    # Return info
    if has_return:
        ws.cell(row=new_excel_row, column=COL_RETDC).value   = str(inv_row[12]).strip()
        ws.cell(row=new_excel_row, column=COL_RETDATE).value = parse_date(inv_row[13])

    # Mark the S.No cell lightly purple so new rows are easy to spot
    ws.cell(row=new_excel_row, column=COL_SN).fill = LIGHT_PURPLE

    print(f"  Added S.No {new_sno}: {serial_str}  |  DC={inv_row[1]}  "
          f"|  Balaji Cost={amt}")


# ═══════════════════════════════════════════════════════════════════════════════
# STEP 6 — Save
# ═══════════════════════════════════════════════════════════════════════════════

print(f"\nSaving → {OUTPUT_PATH} …")
wb.save(OUTPUT_PATH)
print("Done. ✓")
print()
print("Summary")
print("───────")
print(f"  Rows updated  : see purple cells in column W (Balaji Info Cost)")
print(f"  Purple flags  : mismatches in cost, date, and serial number columns")
print(f"  New rows added: S.No {current_max_sno + 1} – "
      f"{current_max_sno + len(rows_to_add)}")
print(f"  Output file   : {OUTPUT_PATH}")
